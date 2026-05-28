"""
Preditor de Deformação Permanente de Solos
DNIT 179/2018-IE | Random Forest + Modelo de Guimarães (2009)
"""

from flask import Flask, render_template, request, jsonify, send_file
import numpy as np
import pickle
import os
import io
import base64
from datetime import datetime
from scipy.optimize import curve_fit
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

app = Flask(__name__)

MODEL_PATH = os.path.join(os.path.dirname(__file__), "data", "model_rf.pkl")
with open(MODEL_PATH, "rb") as f:
    MODEL = pickle.load(f)

P0_ATM = 101.325
CICLOS = [1,10,50,100,500,1000,2000,5000,10000,20000,50000,100000,150000]
CL_LABELS = ["1","10","50","100","500","1k","2k","5k","10k","20k","50k","100k","150k"]

DNIT_TABELA2 = {
    40:  [{"sd":40,"razao":"2,0"},{"sd":80,"razao":"3,0"},{"sd":120,"razao":"4,0"}],
    80:  [{"sd":80,"razao":"2,0"},{"sd":160,"razao":"3,0"},{"sd":240,"razao":"4,0"}],
    120: [{"sd":120,"razao":"2,0"},{"sd":240,"razao":"3,0"},{"sd":360,"razao":"4,0"}],
}

RANGES = {
    "hot":  {"min":8.0,  "max":21.0,  "unit":"%",     "label":"Umidade Ótima (Wot)"},
    "gd":   {"min":1.63, "max":2.11,  "unit":"g/cm³", "label":"γd máx"},
    "p10":  {"min":21.0, "max":100.0, "unit":"%",     "label":"Passante #10"},
    "p40":  {"min":16.0, "max":100.0, "unit":"%",     "label":"Passante #40"},
    "p200": {"min":4.31, "max":98.0,  "unit":"%",     "label":"Passante #200"},
    "cbr":  {"min":2.0,  "max":34.0,  "unit":"%",     "label":"CBR"},
}


def guimaraes_model(X_inp, psi1, psi2, psi3, psi4):
    s3, sd, N = X_inp
    return psi1 * (s3 / P0_ATM) ** psi2 * (sd / P0_ATM) ** psi3 * N ** psi4


def predict_curve(s3_kpa, sd_kpa, hot, gd, p10, p40, p200, cbr):
    s3 = s3_kpa / 1000.0
    sd = sd_kpa / 1000.0
    rows = [[np.log10(c), s3, sd, sd/s3, sd+s3, hot, gd, p10, p40, p200, cbr] for c in CICLOS]
    preds = MODEL.predict(np.array(rows))
    return np.maximum(preds, 0.0)


def fit_guimaraes(dp_curve, s3_kpa, sd_kpa):
    N_arr  = np.array(CICLOS, dtype=float)
    s3_arr = np.full(len(CICLOS), float(s3_kpa))
    sd_arr = np.full(len(CICLOS), float(sd_kpa))
    try:
        popt, _ = curve_fit(guimaraes_model, (s3_arr, sd_arr, N_arr), dp_curve,
                            p0=[0.1,0.5,1.0,0.1], maxfev=20000, bounds=(0,100))
        dp_fit = guimaraes_model((s3_arr, sd_arr, N_arr), *popt)
        ss_res = np.sum((dp_curve - dp_fit)**2)
        ss_tot = np.sum((dp_curve - np.mean(dp_curve))**2)
        r2 = 1 - ss_res/ss_tot if ss_tot > 0 else 0.0
        return [round(float(p),4) for p in popt], round(r2,4)
    except Exception:
        return None, None


def classify_behavior(dp_curve):
    dp150 = dp_curve[-1]
    slope = (dp_curve[-1] - dp_curve[8]) / (150000 - 10000)
    if dp150 < 0.5 and slope < 0.000015: return 1
    if dp150 < 1.5 and slope < 0.000060: return 2
    if dp150 < 3.5: return 3
    return 4


@app.route("/")
def index():
    return render_template("index.html", ranges=RANGES, dnit_tabela2=DNIT_TABELA2)


@app.route("/api/tensoes/<int:s3>")
def get_tensoes(s3):
    pares = DNIT_TABELA2.get(s3)
    if not pares:
        return jsonify({"error":"σ₃ inválido. Use 40, 80 ou 120 kPa."}), 400
    return jsonify(pares)


@app.route("/api/calcular", methods=["POST"])
def calcular():
    data = request.get_json()
    try:
        hot  = float(data["hot"])
        gd   = float(data["gd"])
        p10  = float(data["p10"])
        p40  = float(data["p40"])
        p200 = float(data["p200"])
        cbr  = float(data["cbr"])
        s3   = int(data["s3"])
        sd   = int(data["sd"])
        nc   = int(data.get("nc", 150000))
        nome = str(data.get("nome","Amostra")).strip() or "Amostra"
    except (KeyError,ValueError,TypeError) as e:
        return jsonify({"error":f"Parâmetro inválido: {e}"}), 400

    erros = []
    r = RANGES
    if not (r["hot"]["min"]  <= hot  <= r["hot"]["max"]):
        erros.append(f"Wot ({r['hot']['min']}–{r['hot']['max']}%)")
    if not (r["gd"]["min"]   <= gd   <= r["gd"]["max"]):
        erros.append(f"γd ({r['gd']['min']}–{r['gd']['max']} g/cm³)")
    if not (r["p10"]["min"]  <= p10  <= r["p10"]["max"]):
        erros.append(f"#10 ({r['p10']['min']}–{r['p10']['max']}%)")
    if not (r["p40"]["min"]  <= p40  <= r["p40"]["max"]):
        erros.append(f"#40 ({r['p40']['min']}–{r['p40']['max']}%)")
    if not (r["p200"]["min"] <= p200 <= r["p200"]["max"]):
        erros.append(f"#200 ({r['p200']['min']}–{r['p200']['max']}%)")
    if not (r["cbr"]["min"]  <= cbr  <= r["cbr"]["max"]):
        erros.append(f"CBR ({r['cbr']['min']}–{r['cbr']['max']}%)")
    if erros:
        return jsonify({"error":"Valores fora do intervalo de treino: " + ", ".join(erros)}), 422

    dp_curve = predict_curve(s3, sd, hot, gd, p10, p40, p200, cbr)

    dp_n = dp_curve[-1]
    if nc in CICLOS:
        dp_n = dp_curve[CICLOS.index(nc)]
    else:
        for i in range(len(CICLOS)-1):
            if CICLOS[i] < nc < CICLOS[i+1]:
                rv = (nc-CICLOS[i])/(CICLOS[i+1]-CICLOS[i])
                dp_n = dp_curve[i] + rv*(dp_curve[i+1]-dp_curve[i])
                break

    guim_params, guim_r2 = fit_guimaraes(dp_curve, s3, sd)
    tipo = classify_behavior(dp_curve)

    return jsonify({
        "curva":     dp_curve.round(4).tolist(),
        "ciclos":    CICLOS,
        "labels":    CL_LABELS,
        "dp_n":      round(float(dp_n),4),
        "dp_10k":    round(float(dp_curve[8]),4),
        "dp_50k":    round(float(dp_curve[10]),4),
        "dp_100k":   round(float(dp_curve[11]),4),
        "dp_150k":   round(float(dp_curve[12]),4),
        "tipo":      tipo,
        "guimaraes": guim_params,
        "guim_r2":   guim_r2,
        "nome":      nome,
        "inputs":    {"hot":hot,"gd":gd,"p10":p10,"p40":p40,"p200":p200,
                      "cbr":cbr,"s3":s3,"sd":sd,"nc":nc,"razao":round((s3+sd)/s3,1)},
    })


@app.route("/api/exportar/pdf", methods=["POST"])
def exportar_pdf():
    data     = request.get_json()
    inp      = data.get("inputs", {})
    dp_n     = data.get("dp_n", 0)
    dp_10k   = data.get("dp_10k", 0)
    dp_50k   = data.get("dp_50k", 0)
    dp_100k  = data.get("dp_100k", 0)
    dp_150k  = data.get("dp_150k", 0)
    tipo     = data.get("tipo", 0)
    guim     = data.get("guimaraes") or []
    guim_r2  = data.get("guim_r2")
    nome     = data.get("nome", "Amostra")
    chart_b64= data.get("chart_img", "")

    # ASCII-safe labels (Helvetica nao suporta unicode grego)
    TIPOS_NOME = {
        1: "Tipo I - Acomodamento plastico (rapido)",
        2: "Tipo II - Acomodamento plastico (tardio)",
        3: "Tipo III - Sem acomodamento",
        4: "Tipo IV - Colapso incremental",
    }
    DIAG_LABELS = {
        1: ("Indicado para base", "Indicado para sub-base", "Indicado para subleito"),
        2: ("Indicado com controle tecnico", "Indicado para sub-base", "Indicado para subleito"),
        3: ("Nao indicado para base", "Com restricoes - requer estudo", "Indicado para subleito"),
        4: ("Nao indicado", "Nao indicado", "Usar somente com reforco"),
    }

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def _hdr(cells, widths, fill=(21, 88, 160)):
        pdf.set_fill_color(*fill)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        for txt, w in zip(cells, widths):
            pdf.cell(w, 7, txt, border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    def _row(cells, widths):
        pdf.set_fill_color(245, 245, 245)
        pdf.set_font("Helvetica", "", 9)
        for txt, w in zip(cells, widths):
            pdf.cell(w, 6, str(txt), border=1, align="C", fill=True)
        pdf.ln()

    def _section(title):
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, title, ln=True)
        pdf.set_text_color(0, 0, 0)

    # ── cabeçalho ──
    pdf.set_fill_color(10, 37, 64)
    pdf.rect(0, 0, 210, 28, "F")
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 7)
    pdf.cell(0, 7, "Preditor de Deformacao Permanente - DNIT 179/2018-IE", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(148, 189, 232)
    pdf.set_x(10)
    pdf.cell(0, 6, "Random Forest  |  9.024 ensaios triaxiais  |  15 solos do Ceara", ln=True)
    pdf.set_text_color(0, 0, 0)

    pdf.set_xy(10, 33)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, f"Amostra: {nome}", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 100)
    nc_fmt = f"{inp.get('nc', 0):,}".replace(",", ".")
    pdf.cell(0, 5,
             f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
             f"s3 = {inp.get('s3')} kPa  |  sd = {inp.get('sd')} kPa  |  "
             f"s1/s3 = {inp.get('razao')}  |  N ref. = {nc_fmt}",
             ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── destaque ep ──
    pdf.set_fill_color(10, 37, 64)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"  Deformacao permanente especifica prevista (ep) a {nc_fmt} ciclos",
             fill=True, ln=True)
    pdf.set_font("Helvetica", "B", 28)
    pdf.set_fill_color(21, 88, 160)
    pdf.cell(0, 16, f"  {dp_n:.3f} %", fill=True, ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── métricas ──
    _section("METRICAS DE DEFORMACAO PERMANENTE")
    _hdr(["N = 10.000", "N = 50.000", "N = 100.000", "N = 150.000"], [47, 47, 47, 47])
    _row([f"{dp_10k:.3f}%", f"{dp_50k:.3f}%", f"{dp_100k:.3f}%", f"{dp_150k:.3f}%"],
         [47, 47, 47, 47])
    pdf.ln(4)

    # ── gráfico ──
    if chart_b64:
        try:
            img_data = base64.b64decode(chart_b64.split(",")[-1])
            img_buf = io.BytesIO(img_data)
            _section("CURVA ep x NUMERO DE CICLOS")
            pdf.image(img_buf, x=10, w=188, h=80)
            pdf.ln(4)
        except Exception:
            pass

    # ── tipo ──
    _section("CLASSIFICACAO DO COMPORTAMENTO")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_fill_color(234, 242, 252)
    pdf.cell(0, 8, f"  {TIPOS_NOME.get(tipo, '-')}", fill=True, border=1, ln=True)
    pdf.ln(3)

    # ── diagnóstico ──
    _section("DIAGNOSTICO TECNICO - USO EM PAVIMENTO")
    diag = DIAG_LABELS.get(tipo, ("-", "-", "-"))
    for camada, txt in zip(["Base", "Sub-base", "Subleito"], diag):
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(28, 6, camada + ":", border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(160, 6, txt, border=1, ln=True)
    pdf.ln(4)

    # ── guimarães ──
    if guim:
        _section("MODELO DE GUIMARAES (2009) - ep = psi1*(s3/p0)^psi2 * (sd/p0)^psi3 * N^psi4")
        _hdr(["psi1", "psi2", "psi3", "psi4"], [47, 47, 47, 47])
        _row([str(v) for v in guim], [47, 47, 47, 47])
        if guim_r2 is not None:
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 5, f"R2 do ajuste: {guim_r2}", ln=True)
            pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

    # ── parâmetros do solo ──
    _section("PARAMETROS DO SOLO INSERIDOS")
    params = [
        ("Wot (%)", inp.get("hot")), ("gd max (g/cm3)", inp.get("gd")),
        ("Pass. #10 (%)", inp.get("p10")), ("Pass. #40 (%)", inp.get("p40")),
        ("Pass. #200 (%)", inp.get("p200")), ("CBR (%)", inp.get("cbr")),
    ]
    _hdr([p[0] for p in params], [31] * 6, fill=(10, 37, 64))
    _row([str(p[1]) for p in params], [31] * 6)
    pdf.ln(5)

    # ── rodapé ──
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5,
             "DNIT 179/2018-IE - Pavimentacao - Solos - Deformacao permanente  |  "
             "Guimaraes (2009) COPPE/UFRJ",
             ln=True)

    buf = io.BytesIO(bytes(pdf.output()))
    buf.seek(0)
    safe = nome.replace(" ", "_").replace("/", "-")
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=f"DP_{safe}.pdf")


@app.route("/api/exportar/xlsx", methods=["POST"])
def exportar_xlsx():
    data    = request.get_json()
    inp     = data.get("inputs", {})
    curva   = data.get("curva", [])
    ciclos  = data.get("ciclos", [])
    labels  = data.get("labels", [])
    guim    = data.get("guimaraes") or []
    guim_r2 = data.get("guim_r2")
    nome    = data.get("nome", "Amostra")
    tipo    = data.get("tipo", 0)

    TIPOS_NOME = {1:"Tipo I — Acomodamento plástico (rápido)",
                  2:"Tipo II — Acomodamento plástico (tardio)",
                  3:"Tipo III — Sem acomodamento",
                  4:"Tipo IV — Colapso incremental"}

    wb = openpyxl.Workbook()

    # ── Aba 1: Resumo ──
    ws = wb.active
    ws.title = "Resumo"
    navy = "0A2540"
    blue = "1558A0"
    blue_l = "EAF2FC"
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill_navy = PatternFill("solid", fgColor=navy)
    hdr_fill_blue = PatternFill("solid", fgColor=blue)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def hdr_cell(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font
        c.fill = fill or hdr_fill_blue
        c.alignment = center
        c.border = thin
        return c

    def data_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = center
        c.border = thin
        c.fill = PatternFill("solid", fgColor="F5F5F5")
        return c

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Preditor de Deformação Permanente — DNIT 179/2018-IE"
    t.font = Font(bold=True, color="FFFFFF", size=12)
    t.fill = hdr_fill_navy
    t.alignment = center
    ws.row_dimensions[1].height = 22

    ws["A2"] = f"Amostra: {nome}"
    ws["A2"].font = Font(bold=True, size=10)
    ws["D2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["D2"].font = Font(size=9, color="666666")

    # tensões
    r = 4
    for lbl, val in [("σ₃ (kPa)", inp.get("s3")), ("σd (kPa)", inp.get("sd")),
                     ("σ₁/σ₃", inp.get("razao")), ("N ref.", inp.get("nc"))]:
        hdr_cell(ws, r, 1, lbl, hdr_fill_navy)
        data_cell(ws, r, 2, val)
        r += 1

    # métricas εp
    r = 4
    for lbl, val in [("εp a N ref. (%)", data.get("dp_n")), ("εp a 10.000 (%)", data.get("dp_10k")),
                     ("εp a 50.000 (%)", data.get("dp_50k")), ("εp a 100.000 (%)", data.get("dp_100k")),
                     ("εp a 150.000 (%)", data.get("dp_150k"))]:
        hdr_cell(ws, r, 4, lbl)
        data_cell(ws, r, 5, val)
        r += 1

    # tipo
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value=f"Comportamento: {TIPOS_NOME.get(tipo,'—')}")
    c.font = Font(bold=True, size=10, color=navy)
    c.fill = PatternFill("solid", fgColor=blue_l)
    c.alignment = center
    c.border = thin

    # parâmetros do solo
    r += 2
    solo_params = [("Wot (%)", inp.get("hot")), ("γd máx (g/cm³)", inp.get("gd")),
                   ("Pass. #10 (%)", inp.get("p10")), ("Pass. #40 (%)", inp.get("p40")),
                   ("Pass. #200 (%)", inp.get("p200")), ("CBR (%)", inp.get("cbr"))]
    for i, (lbl, val) in enumerate(solo_params):
        hdr_cell(ws, r, i+1, lbl, hdr_fill_navy)
        data_cell(ws, r+1, i+1, val)
    ws.row_dimensions[r].height = 18

    # guimarães
    if guim:
        r += 3
        for i, sym in enumerate(["ψ₁","ψ₂","ψ₃","ψ₄"]):
            hdr_cell(ws, r, i+1, sym)
            data_cell(ws, r+1, i+1, guim[i])
        if guim_r2 is not None:
            ws.cell(row=r+2, column=1, value=f"R² do ajuste: {guim_r2}").font = Font(size=9, italic=True)

    from openpyxl.utils import get_column_letter
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 22

    # ── Aba 2: Curva εp ──
    ws2 = wb.create_sheet("Curva εp")
    hdr_cell(ws2, 1, 1, "N (ciclos)", hdr_fill_navy)
    hdr_cell(ws2, 1, 2, "εp (%)", hdr_fill_navy)
    hdr_cell(ws2, 1, 3, "Label", hdr_fill_navy)
    for i, (n, ep, lbl) in enumerate(zip(ciclos, curva, labels), start=2):
        data_cell(ws2, i, 1, n)
        data_cell(ws2, i, 2, ep)
        data_cell(ws2, i, 3, lbl)

    chart = LineChart()
    chart.title = f"Curva εp — {nome}"
    chart.style = 10
    chart.y_axis.title = "εp (%)"
    chart.x_axis.title = "Ponto de leitura"
    chart.height = 12
    chart.width = 20
    ep_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(ciclos)+1)
    chart.add_data(ep_ref, titles_from_data=True)
    ws2.add_chart(chart, "E2")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = nome.replace(" ","_").replace("/","-")
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"DP_{safe}.xlsx")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
